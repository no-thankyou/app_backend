"""Классы для настройки админки."""
import openpyxl
from django.contrib import admin
from django.contrib.auth.admin import UserAdmin
from django.forms import forms
from django.shortcuts import redirect, render
from django.urls import path
from django.utils.translation import gettext_lazy as _

from api.models import (City, Competence, Event, EventPhoto, Tags, User,
                        SMSAuth, Participation)


class FileForm(forms.Form):
    """Форма для получения файла с пользователями."""

    xlsx_file = forms.FileField()


@admin.register(User)
class CustomUserAdmin(UserAdmin):
    """Класс для настроек админки модели Пользователь."""

    change_list_template = 'admin/import_users_changelist.html'
    fieldsets = (
        (None, {'fields': ('username', 'password')}),
        (_('Personal info'), {'fields': ('first_name', 'last_name', 'email',
                                         'phone')}),
        ('Информация', {'fields': ('city', 'profession', 'competences',
                                        'photo')}),
        ('Контакты', {'fields': ('about', 'website', 'telegram',
                                 'instagram')}),
        (_('Permissions'), {
            'fields': ('is_active', 'is_staff', 'is_superuser', 'groups',
                       'user_permissions'),
        }),
        (_('Important dates'), {'fields': ('last_login', 'date_joined',
                                           'subscription_expiration_date',
                                           'date_of_registration')}))
    list_display = ('phone', 'subscription_expiration_date', 'first_name',
                    'last_name', 'is_superuser')
    list_filter = ('subscription_expiration_date', 'is_superuser')
    filter_horizontal = ('competences', 'groups',
                         'user_permissions')
    search_fields = ('phone', 'first_name', 'last_name')

    def get_urls(self):
        """Добавление кастомных запросов в админку."""
        return [
            path('import-users/', self.import_users, name='import-users'),
        ] + super().get_urls()

    def import_users(self, request):
        """Обработчик загрузки пользователей из файла."""
        if request.method == 'POST':
            book = openpyxl.load_workbook(request.FILES['xlsx_file'],
                                          read_only=True, data_only=True)
            sheet = book.get_sheet_by_name(book.get_sheet_names()[0])
            head = [cell.value for cell in sheet['A1':'D1'][0]]

            for row in sheet['A2':f'D{sheet.max_row}']:
                item = self._prepare_row(head, row)
                user, _ = User.objects.get_or_create(phone=item['Телефон'],
                                                     username=item['Телефон'])
                user.last_name = item['Фамилия']
                user.first_name = item['Имя']
                user.subscription_expiration_date = item['Подписка']
                user.save()
            self.message_user(request, 'Пользователи загружены')
            return redirect('..')
        return render(request, 'admin/import_file_form.html',
                      {'form': FileForm()})

    def _prepare_row(self, head, row):
        """Обработка строки с пользователем из файла."""
        item = {head[cell.column - 1]: cell.value for cell in row}
        item['Телефон'] = str(item['Телефон'])
        if not item['Телефон'].startswith('+'):
            item['Телефон'] = '+' + item['Телефон']
        return item

    def save_model(self, request, obj, form, change):
        """Замена пустого поля phone номером телефона."""
        # в username летит номер телефона, особенность регистрации
        if not obj.phone:
            obj.phone = obj.username
        return super().save_model(request, obj, form, change)


class EventPhotoInline(admin.TabularInline):
    """Класс настройки отображения фото события в админке."""

    model = EventPhoto


class ParticipationInline(admin.StackedInline):
    """Класс настройки участия в событии."""

    model = Participation
    exclude = ('user', )
    extra = 0


@admin.register(Event)
class CustomEventAdmin(admin.ModelAdmin):
    """Класс для настроек админки модели События."""

    fields = ('title', 'description', 'start_date', 'end_date',
              'address', 'tags')
    list_display = ('title', 'start_date', 'end_date', 'address')
    list_filter = ('start_date',)
    search_fields = ['title', 'address']
    filter_horizontal = ('tags',)
    inlines = [EventPhotoInline, ParticipationInline]
    change_list_template = 'admin/import_events_changelist.html'

    def get_urls(self):
        """Добавление кастомных запросов в админку."""
        return [
            path('import-events/', self.import_events, name='import-events'),
        ] + super().get_urls()

    def import_events(self, request):
        """Обработчик загрузки событий из файла."""
        if request.method == 'POST':
            book = openpyxl.load_workbook(request.FILES['xlsx_file'],
                                          data_only=True)
            sheet = book.get_sheet_by_name(book.get_sheet_names()[0])
            head = [cell.value for cell in sheet['A1':'D1'][0]]
            for row in sheet['A2':'D26']:
                item = self._prepare_row(head, row)
                event, _ = (Event.objects.get_or_create
                            (title=item['Название'],
                             description=item['Описание'],
                             start_date=item['Дата начала события'],
                             end_date=item['Дата окончания события']))

            return redirect('..')
        return render(request, 'admin/import_file_form.html',
                      {'form': FileForm()})

    def _prepare_row(self, head, row):
        item = {head[cell.column - 1]: cell.value for cell in row}
        if item['Описание'] is None:
            item['Описание'] = ''
        return item


@admin.register(Tags)
class CustomTagsAdmin(admin.ModelAdmin):
    """Класс для настроек админки модели Теги."""

    list_display = ('name', 'city')
    list_filter = ('name',)
    search_fields = ('name',)


@admin.register(City)
class CustomCityAdmin(admin.ModelAdmin):
    """Класс для настроек админки модели Города."""

    ordering = ('name', )
    search_fields = ('name',)
    change_list_template = 'admin/import_cities_changelist.html'

    def get_urls(self):
        """Добавление кастомных запросов в админку."""
        return [
            path('import-cities/', self.import_cities,
                 name='import-cities'),
        ] + super().get_urls()

    def import_cities(self, request):
        """Обработчик загрузки городов из файла."""
        if request.method == 'POST':
            uploaded_file = request.FILES['xlsx_file']
            cities = uploaded_file.read().decode('utf-8').splitlines()

            for city in cities:
                City.objects.get_or_create(name=city)
            self.message_user(request, 'Города загружены')
            return redirect('..')
        return render(request, 'admin/import_file_form.html',
                      {'form': FileForm()})


@admin.register(Competence)
class CustomCompetenceAdmin(admin.ModelAdmin):
    """Класс для настроек админки модели Компетенции."""

    search_fields = ('name',)
    change_list_template = 'admin/import_competences_changelist.html'

    def get_urls(self):
        """Добавление кастомных запросов в админку."""
        return [
            path('import-competences/', self.import_competences,
                 name='import-competences'),
        ] + super().get_urls()

    def import_competences(self, request):
        """Обработчик загрузки компетенций из файла."""
        if request.method == 'POST':
            book = openpyxl.load_workbook(request.FILES['xlsx_file'],
                                          read_only=True, data_only=True)
            sheet = book.get_sheet_by_name(book.get_sheet_names()[0])
            for row in sheet['A1':f'A{sheet.max_row}']:
                for name in row:
                    Competence.objects.get_or_create(name=name.value)
            self.message_user(request, 'Компетенции загружены')
            return redirect('..')
        return render(request, 'admin/import_file_form.html',
                      {'form': FileForm()})


@admin.register(SMSAuth)
class CustomSMSAuthAdmin(admin.ModelAdmin):
    """Класс для настроек админки модели Смс."""

    list_display = ('phone', 'code', 'attempts', 'send_at', 'is_used')
